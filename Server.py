#!/usr/bin/env python
#coding:utf-8

import textwrap
import json
import tornado.httpserver
import tornado.ioloop
import tornado.options
import tornado.web

import http.server
import socketserver
import os
import time

from openpyxl import load_workbook
from math import radians, cos, sin, asin, sqrt,fabs
from dictionary import ROAD_NAME,JTJ,distance_graph,predict
import copy

Time=time.strftime("%H:%M", time.localtime())
Handler = http.server.SimpleHTTPRequestHandler

from tornado.options import define, options
define("port", default=8000, help="Please send email to me", type=int)

class ReverseHandler(tornado.web.RequestHandler):
	def get(self, input_word):##input=http://localhost:8000/navigation/start="str"&end="str"&transpotation="str"
		global Time
		index=0
		start=input_word[input_word.find('=',index)+1:input_word.find('&',index)]
		index=input_word.find('&',index)+1		
		end=input_word[input_word.find('=',index)+1:input_word.find('&',index)]
		index=input_word.find('&',index)
		transpotation=input_word[input_word.find('=',index)+1::]
		
		fpcomma=start.find(',')
		tpcomma=end.find(',')
		fp=[float(start[:fpcomma]),float(start[fpcomma+1::])]
		tp=[float(end[:tpcomma]),float(end[tpcomma+1::])]
		if transpotation=="scooter":
			motorcycle=True
		else:
			motorcycle=False

		def haversine(lon1, lat1, lon2, lat2):          # this is a formula that use GPS to compute the distance
			lon1, lat1, lon2, lat2 = map(radians, [lon1, lat1, lon2, lat2])
			dlon = lon2 - lon1
			dlat = lat2 - lat1
			a = sin(dlat/2)**2 + cos(lat1) * cos(lat2) * sin(dlon/2)**2
			c = 2 * asin(sqrt(a))
			r = 6371
			return c * r * 1000

		def dijkstra(graph,start,goal):                 # this is an agorithm used to compute the smallest time
			shortest_time = {}
			predecessor = {}
			unseenNodes = graph
			infinity = 9999999
			for node in unseenNodes:
				shortest_time[node] = infinity
			shortest_time[start] = 0
 
			while unseenNodes:
				minNode = None
				for node in unseenNodes:
					if minNode is None:
						minNode = node
					elif shortest_time[node] < shortest_time[minNode]:
						minNode = node
 
				for childNode, weight in graph[minNode].items():
					if weight + shortest_time[minNode] < shortest_time[childNode]:
						shortest_time[childNode] = weight + shortest_time[minNode]
						predecessor[childNode] = minNode
				unseenNodes.pop(minNode)
 
			currentNode = goal
			while currentNode != start:
				try:
					Juction.insert(0,currentNode)
					currentNode = predecessor[currentNode]
				except KeyError:
					print('Juction not reachable')
					break
			Juction.insert(0,start)
			if shortest_time[goal] != infinity:
				return shortest_time[goal]+ft+tt+len(Juction)*27

			
		road_to_juction = load_workbook("RTJ.xlsx")     # The file road can to what juction
		juction = load_workbook("juction_data.xlsx")    # The file the juction id and its GPS
		deactive = load_workbook("deactive.xlsx")       # The file the government gave us the roadname and its GPS
		direction = load_workbook("direction.xlsx")     # The file used to find the last juction

		JD = juction.worksheets[0]
		RTJ = road_to_juction.worksheets[0]
		deactivesheet = deactive.worksheets[0]
		DIR = direction.worksheets[0]
		DIR = direction.worksheets[0]

		###draw the graph
		Juction=[]                                      # it will record all juction we passed
		time_graph=copy.deepcopy(JTJ)                   # time graph is record how many times juction to juction
		time_graph2=copy.deepcopy(JTJ)                  # time graph is record how many times juction to juction
		all_distance=0                                  
		
		while True:
			if Time not in predict:
				lastnumber=ord(Time[len(Time)-1])+1
				if lastnumber > 57:
					lastnumber=48
				Time=Time[:4]+chr(lastnumber)
			else:
				break
		
		for Start in JTJ:
			for End in JTJ[Start]:
				try:
					roadname=JTJ[Start][End]
					speed=predict[Time][roadname]
					distance=distance_graph[Start][End];
					time=distance/speed
					time_graph[Start][End]=time
					time_graph2[Start][End]=time
				except:
					print(roadname)
					print(time)
		
		fptofr=99999                                    # default the distance between the point and road
		tptotr=99999
		
		def length(x,y):
			return sqrt(x*x+y*y)

		for i in range(2,deactivesheet.max_row+1):

			line_x1=deactivesheet["B"+str(i)].value   # I use the formula a dot to a line segment 
			line_y1=deactivesheet["A"+str(i)].value        
			line_x2=deactivesheet["D"+str(i)].value    
			line_y2=deactivesheet["C"+str(i)].value        
	                                          
			vector_vx = line_x2-line_x1;
			vector_vy = line_y2-line_y1;
	
			vector_fpvx1 = fp[0] - line_x1;
			vector_fpvy1 = fp[1] - line_y1;
			vector_fpvx2 = fp[0] - line_x2;
			vector_fpvy2 = fp[1] - line_y2;
	
			vector_tpvx1 = tp[0] - line_x1;
			vector_tpvy1 = tp[1] - line_y1;
			vector_tpvx2 = tp[0] - line_x2;
			vector_tpvy2 = tp[1] - line_y2;
	
			rfp=(vector_fpvx1*vector_vx+vector_fpvy1*vector_vy)/(length(vector_vx,vector_vy)*length(vector_vx,vector_vy))
	
			if(rfp<=0):
				if(length(vector_fpvx1,vector_fpvy1)<fptofr):
					fptofr=length(vector_fpvx1,vector_fpvy1)
					fr=deactivesheet["F"+str(i)].value
			elif(rfp>=1):
				if(length(vector_fpvx2,vector_fpvy2)<fptofr):
					fptofr=length(vector_fpvx2,vector_fpvy2)
					fr=deactivesheet["F"+str(i)].value
			else:
				if((fabs(vector_vx*vector_fpvy1-vector_fpvx1*vector_vy)/length(vector_vx,vector_vy))<fptofr):
					fptofr=fabs(vector_vx*vector_fpvy1-vector_fpvx1*vector_vy)/length(vector_vx,vector_vy)
					fr=deactivesheet["F"+str(i)].value

			rtp=(vector_tpvx1*vector_vx+vector_tpvy1*vector_vy)/(length(vector_vx,vector_vy)*length(vector_vx,vector_vy))
	
			if(rtp<=0):
				if(length(vector_tpvx1,vector_tpvy1)<tptotr):
					tptotr=length(vector_tpvx1,vector_tpvy1)
					tr=deactivesheet["F"+str(i)].value
			elif(rtp>=1):
				if(length(vector_tpvx2,vector_tpvy2)<tptotr):
					tptotr=length(vector_tpvx2,vector_tpvy2)
					tr=deactivesheet["F"+str(i)].value
			else:
				if((fabs(vector_vx*vector_tpvy1-vector_tpvx1*vector_vy)/length(vector_vx,vector_vy))<tptotr):
					tptotr=fabs(vector_vx*vector_tpvy1-vector_tpvx1*vector_vy)/length(vector_vx,vector_vy)
					tr=deactivesheet["F"+str(i)].value
		
		for i in range(2,RTJ.max_row+1):                # when i got the fr and tr
			if RTJ["A"+str(i)].value==fr:               # i need to konw  the first juctionandthe lst juction
				fj=RTJ["B"+str(i)].value                # because the time graph is used juction to be the refernce.
			if DIR["A"+str(i)].value==tr:
				for j in range(2,JD.max_row+1):
					if DIR["C"+str(i)].value==JD["B"+str(j)].value:
						tj=JD["A"+str(j)].value
		
		fs=predict[Time][fr]
		ts=predict[Time][tr]
		ft=haversine(fp[1],fp[0],JD["C"+str(fj+1)].value,JD["D"+str(fj+1)].value)/fs+haversine(fp[1],fp[0],JD["C"+str(fj+1)].value,JD["D"+str(fj+1)].value)*0.15
		tt=haversine(tp[1],tp[0],JD["C"+str(tj+1)].value,JD["D"+str(tj+1)].value)/ts+haversine(tp[1],tp[0],JD["C"+str(tj+1)].value,JD["D"+str(tj+1)].value)*0.15
	
		all_time=dijkstra(time_graph,fj,tj)            # use dijkstra to compute what road is smallest time
		
		motorcycle_turn=0
		#print("1:\t",fr,"\t",int(ft),"s\t",int(haversine(fp[1],fp[0],JD["C"+str(fj+1)].value,JD["D"+str(fj+1)].value)),"m\t",ROAD_NAME[fr],"\t")
		all_distance+=haversine(fp[1],fp[0],JD["C"+str(fj+1)].value,JD["D"+str(fj+1)].value)
		now=ROAD_NAME[fr]
		for i in range(0,len(Juction)-1):
			#print(i+2,":\t",JTJ[Juction[i]][Juction[i+1]],"\t",int(time_graph2[Juction[i]][Juction[i+1]]),"s\t",int(distance_graph[Juction[i]][Juction[i+1]]),"m\t",ROAD_NAME[JTJ[Juction[i]][Juction[i+1]]],"\t")
			if now[:2]!=ROAD_NAME[JTJ[Juction[i]][Juction[i+1]]][:2]:
				motorcycle_turn+=1
				now=ROAD_NAME[JTJ[Juction[i]][Juction[i+1]]][:2]
			all_distance+=distance_graph[Juction[i]][Juction[i+1]]
		else:
			#print(len(Juction)+1,":\t",tr,"\t",int(tt),"s\t",int(haversine(tp[1],tp[0],JD["C"+str(tj+1)].value,JD["D"+str(tj+1)].value)),"m\t",ROAD_NAME[tr],"\t")
			if now[:2]!=ROAD_NAME[tr][:2]:
				motorcycle_turn+=1
		all_distance+=haversine(tp[1],tp[0],JD["C"+str(tj+1)].value,JD["D"+str(tj+1)].value)
		if motorcycle:
			all_time=all_time+27*motorcycle_turn
		
		Passed=[]
		output={"time":all_time,"distance":all_distance,"start":[fp[1],fp[0]],"end":[tp[1],tp[0]],"waypoint":{}}
		###output={"start":[lat,lng],"end":[lat,lng],"waypoint":{1:[lat,lng],2:[lat,lng]}}
		index=1
		for i in Juction:
			output["waypoint"].update({index:[JD["C"+str(i+1)].value,JD["D"+str(i+1)].value]})
			Passed.append([JD["C"+str(i+1)].value,JD["D"+str(i+1)].value])
			index+=1
		[v for v in sorted(output["waypoint"].values())]
		
		self.write(output)




if __name__ == "__main__":
    tornado.options.parse_command_line()
    app = tornado.web.Application(
        handlers = [(r"/navigation/(.*)", ReverseHandler)]
    )
    http_server = tornado.httpserver.HTTPServer(app)
    http_server.listen(options.port)
    tornado.ioloop.IOLoop.instance().start()